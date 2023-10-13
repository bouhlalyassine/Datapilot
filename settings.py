import json
from pathlib import Path
import tempfile
import requests
import openpyxl as oxl
import pandas as pd
from fpdf import FPDF
import streamlit as st
import plotly.graph_objects as go
from openpyxl.styles import Font, Color, Fill, PatternFill, Border, Side, numbers, Alignment
from openpyxl.utils import get_column_letter

TITLE = "Datapilot"


current_dir = Path(__file__).parent if "__file__" in locals() else Path.cwd()


sql_db_p = current_dir / "files" / "SQL_DB.db"
sql_db = str(sql_db_p)


# My Tuto :
space = 15
tuto_space = 70

tuto_sql_db_p = current_dir / "files" / "tuto_sql_db.mp4"
tuto_sql_db = str(tuto_sql_db_p)

lottie_sql_db = current_dir / "files" / "sql_db.json"


def load_lottiefile(filepath : str):
    with open(filepath, "r") as f:
        return json.load(f)


def load_lottieurl(url: str):
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()


@st.cache_data
def curstom_excel_df(excel_file):
    wb = oxl.load_workbook(excel_file)

    # Backgrounds
    Gris_fonce = PatternFill(patternType='solid', fgColor="0A0A0A") # Gris foncé
    Marron_Clair = PatternFill(patternType='solid', fgColor="C4BD97") # Marron clair
    Orange_Clair = PatternFill(patternType='solid', fgColor="FABF8F") # Orange clair
    Vert_fonce = PatternFill(patternType='solid', fgColor="3BCD40") # Vert Foncé
    Vert_Clair = PatternFill(patternType='solid', fgColor="92D050") # Vert clair
    Rouge_Rose = PatternFill(patternType='solid', fgColor="FF7C80") # Rouge/Rose 

    # Borders
    side1 = Side(border_style='thin', color="000000")
    border1 = Border(top=side1 , bottom=side1, right=side1, left=side1)

    # Fontstyles
    font_style = Font(name= 'Calibri', size = 11, color= "000000", bold = True)

    # Alignements
    alignment_left = Alignment(horizontal="left",vertical="center")
    alignment_center = Alignment(horizontal="center",vertical="center")


    """-----------------------------Mise en forme Feuilles -----------------------------""" 
    ws = wb[wb.sheetnames[0]]
    max_c = ws.max_column # le max de cols (exploités) dans la feuille
    max_r = ws.max_row # le max de lignes (exploités) dans la feuille

    cell_fil1 = get_column_letter(max_c)
    ws.auto_filter.ref = f'A1:{cell_fil1}1' # Add a filter
    
    # Mise en forme cadre du tableau en gris :
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=max_r+10, max_col = max_c+10):
        for cell in row:
            cell.fill=Gris_fonce

    # Mise en forme en-tete :
    for cell in ws[1]:
        cell.font=font_style

    # Hauteur de la 1ere ligne
    ws.row_dimensions[1].height = 30

    for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col = max_c):
        for cell in row:
            cell.fill=Marron_Clair
            cell.border=border1
            cell.alignment = alignment_left


    for index_col in range(1, max_c+1):
        i = get_column_letter(index_col)
        ws.column_dimensions[i].width = 13

    wb.save(excel_file)
    
    return excel_file


@st.cache_data
def get_sql_df(df):
    df['total_sale_price'] = df['product_quantity'] * df['product_unit_sale_price']
    df['total_buy_price'] = df['product_quantity'] * df['product_unit_buy_price']

    df['5_percent_taxe'] = df['total_sale_price'] * 0.05

    df['net_profit'] = df['total_sale_price'] - df['total_buy_price'] - df['5_percent_taxe']

    df['sale_date'] = pd.to_datetime(df['sale_date'])

    df['sale_week'] = df['sale_date'].dt.strftime("%U")

    df = df.reindex(columns=["sale_date", 'sale_week', 'customer_name', 'product_name', 
        "product_quantity", 'product_unit_price','total_buy_price','total_sale_price',
        '5_percent_taxe', 'net_profit']).round(2)

    df.sort_values(["sale_date"], ascending=True, inplace=True)

    return df


@st.cache_data
def total_sql_kpi(df):
    tot_sales = round(df['total_sale_price'].sum(),1)
    tot_taxes = round(df['5_percent_taxe'].sum(),1)
    tot_net_profit = round(df['net_profit'].sum(),1)
    tot_rate_profit = round((tot_net_profit / tot_sales) * 100,1)

    return [tot_sales, tot_taxes, tot_net_profit, tot_rate_profit]



@st.cache_data
def month_chart(df):
    df = df.set_index('sale_date')
    df_monthly = round(df.resample('M').sum(),2)
    df_monthly['month_name'] = df_monthly.index.strftime('%B') # %b pour jan au lieu de January

    df_monthly['percent_net_profit'] = round((df_monthly['net_profit'] / df_monthly['total_sale_price'])*100,2)

    # Create the figure
    trace1_0 = go.Bar(x=df_monthly['month_name'], y=df_monthly['total_sale_price'], name="Total Sales",
        marker=dict(color='#2B3DD1'),text=df_monthly['total_sale_price'].apply(lambda x: round(x, 0)),
        textposition='auto',) # .apply(lambda x: '{:.0f}'.format(x))

    trace2_1 = go.Scatter(x=df_monthly['month_name'], y=df_monthly['percent_net_profit'],
        name="% Net Profit", yaxis='y2', line=dict(width=3, color='#44A02D'),
        marker=dict(size=10, color='#44A02D')) # .apply(lambda x: '{:.0f}'.format(x))

    fig = go.Figure()

    fig.add_traces([trace1_0, trace2_1])

    # Update the layout to include two y-axes
    fig.update_layout(
        barmode='group', # stack pour empilé
        yaxis=dict(title="Amount ($)"),

        yaxis2=dict(title="Profit Rate", overlaying='y', side='right',),
        yaxis2_ticksuffix='%',

        font=dict(size=12, family='Arial Black'),
        legend=dict(orientation="h", x=0.1, y=1.2,
                    font=dict(size=12,family='Arial Black')),
        title={
        'text': "Total Sales/Net Profit by Month",
        'y':0.95,
        'x':0.5,
        'xanchor': 'center',
        'yanchor': 'top'
        },
        )
    
    return fig


@st.cache_data
def product_chart(df):
    df_chart = df.groupby(["product_name"], as_index=False).agg({
            'total_sale_price' : 'sum',
            'net_profit' : 'sum',
            }).round(2)

    df_chart['percent_net_profit'] = (df_chart['net_profit'] / df_chart['total_sale_price'])*100

    df_chart.sort_values(["total_sale_price"], ascending=False, inplace=True)

    # Create the figure
    trace1_0 = go.Bar(x=df_chart['product_name'], y=df_chart['total_sale_price'], name="Total Sales",
        marker=dict(color='#2B3DD1'),text=df_chart['total_sale_price'].apply(lambda x: round(x, 0)),
        textposition='auto',width=0.3, offset=-0.15) # .apply(lambda x: '{:.0f}'.format(x))

    trace2_1 = go.Bar(x=df_chart['product_name'], y=df_chart['percent_net_profit'], name="% Net Profit",
        yaxis='y2', marker=dict(color='#44A02D'),
        text=df_chart['percent_net_profit'].apply(lambda x: '{:.0f}%'.format(x)),
        textposition='auto', width=0.3, offset=0.15) # .apply(lambda x: round(x, 0)

    fig = go.Figure()

    fig.add_traces([trace1_0, trace2_1])

    # Update the layout to include two y-axes
    fig.update_layout(
        barmode='group', # stack pour empilé
        yaxis=dict(title="Amount ($)"),
        yaxis2=dict(title="Profit Rate", overlaying='y', side='right',),

        yaxis2_ticksuffix='%',

        font=dict(size=12, family='Arial Black'),
        legend=dict(orientation="h", x=0.1, y=1.2,
                    font=dict(size=12,family='Arial Black')),
        title={
        'text': "Total Sales / % Net Profit by Product",
        'y':0.95,
        'x':0.5,
        'xanchor': 'center',
        'yanchor': 'top'
        },
        )
    
    return fig


@st.cache_data
def customer_chart(df):
    df_chart = df.groupby(["customer_name"], as_index=False).agg({
        'total_sale_price' : 'sum',
        'net_profit' : 'sum',
        }).round(2)
    
    df_chart.sort_values(["total_sale_price"], ascending=False, inplace=True)

    # Create the figure
    trace1_0 = go.Bar(x=df_chart['customer_name'], y=df_chart['total_sale_price'], name="Total Sales",
        marker=dict(color='#2B3DD1'),text=df_chart['total_sale_price'].apply(lambda x: round(x, 0)),
        textposition='auto',) # .apply(lambda x: '{:.0f}'.format(x))

    trace2_1 = go.Bar(x=df_chart['customer_name'], y=df_chart['net_profit'], name="Net Profit",
        marker=dict(color='#44A02D'),text=df_chart['net_profit'].apply(lambda x: round(x, 0)),
        textposition='auto',) # .apply(lambda x: '{:.0f}'.format(x))

    fig = go.Figure()

    fig.add_traces([trace1_0, trace2_1])

    # Update the layout to include two y-axes
    fig.update_layout(
        barmode='group', # stack pour empilé
        yaxis=dict(title="Amount ($)"),
        font=dict(size=12, family='Arial Black'),
        legend=dict(orientation="h", x=0.1, y=1.2,
                    font=dict(size=12,family='Arial Black')),
        title={
        'text': "Total Sales/Net Profit by Cutomer",
        'y':0.95,
        'x':0.5,
        'xanchor': 'center',
        'yanchor': 'top'
        },
        )
    
    return fig



def create_PDF(figs, sales, taxes, profit_tot, profit_percent ):
    # Define the font color as RGB values (dark gray)
    font_color = (0, 0, 0)

    # Create a PDF document and set the page size
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 24)

    # Add the overall page title
    title = "Sales Report" # as of {date.today().strftime('%m/%d/%Y')}
    pdf.set_text_color(*font_color)
    pdf.cell(0, 20, title, align='C', ln=1)

    # Set the line color
    pdf.set_draw_color(0, 0, 0)  # black
    # Draw a horizontal line
    pdf.line(10, pdf.y + 5, pdf.w - 10, pdf.y + 5)
    pdf.ln(5)

    # Add a subheader with KPIs
    pdf.set_font('Arial', '', 16)
    pdf.cell(0, 10, "Total Sales     Total Taxes (5%)     Total Net Profit     Net Profit", align='C')
    pdf.ln(10)
    pdf.cell(0, 10, f"{sales} $            {taxes} $                   {profit_tot} $             {profit_percent}%", align='C')

    #pdf.cell(pdf.w/2, 10, "$100,000", align='L')
    #pdf.ln(5)
    #pdf.cell(pdf.w/2, 10, "Total Units Sold:", align='R')
    #pdf.cell(pdf.w/2, 10, "1,000", align='L')
    
    pdf.ln(5)    
    pdf.line(10, pdf.y + 5, pdf.w - 10, pdf.y + 5)
    pdf.ln(10)

    # Add each chart to the PDF document
    for fig in figs:

        # save the chart as a PNG to a temporary file
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_chart_file:
            fig.write_image(temp_chart_file.name, width=800, height=400)

        # add the chart to the PDF document
        #pdf.ln(5)  # Add padding at the top of the next chart

        chart_width = 165  # set the chart width
        chart_height = 72  # set the chart height
        x_position = (pdf.w - chart_width) / 2  # calculate the x position to center the chart horizontally

        pdf.image(temp_chart_file.name, x=x_position, y=None, w=chart_width, h=chart_height)

    temp = tempfile.NamedTemporaryFile(delete=True)
    pdf_filename = temp.name + f'.pdf'

    pdf.output(pdf_filename, "F")

    # convert the PDF to a base64 string
    with open(pdf_filename, "rb") as f:
        pdf_bytes = f.read()
    
    return pdf_bytes 

